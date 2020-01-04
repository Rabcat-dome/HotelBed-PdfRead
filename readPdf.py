import os
from datetime import datetime
from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.converter import TextConverter
from pdfminer.layout import LAParams
from pdfminer.pdfpage import PDFPage
from io import BytesIO as StringIO
from cStringIO import StringIO as StringIO2
import PyPDF2 

import openpyxl
from openpyxl import Workbook 
from openpyxl import load_workbook
wb = load_workbook('DatabaseHB.xlsm',keep_vba=True)
ws = wb['raw']

path = os.path.join(os.path.dirname(os.path.realpath(__file__)),'pdfFile')
#fileDb = os.path.join(os.path.dirname(os.path.realpath(__file__)),'DatabaseHB.xlsm')
folders = []

# r=root, d=directories, f = files
for r, d, f in os.walk(path):
    for folder in f:
        if folder.endswith('.pdf') or folder.endswith('.PDF'):
            folders.append(os.path.join(r, folder))

def convert_pdf_to_txt(path):
    rsrcmgr = PDFResourceManager()
    retstr = StringIO()
    codec = 'utf-8'
    laparams = LAParams()
    device = TextConverter(rsrcmgr, retstr, codec=codec, laparams=laparams)
    fp = open(path, 'rb')
    interpreter = PDFPageInterpreter(rsrcmgr, device)
    password = ""
    maxpages = 0
    caching = True
    pagenos=set()
    try:
        for page in PDFPage.get_pages(fp, pagenos, maxpages=maxpages, password=password,caching=caching, check_extractable=True):
            interpreter.process_page(page)
        text = retstr.getvalue()
        
    except:
        print("Mal PDF FORM") 
    
    fp.close()
    device.close()
    retstr.close()

    startT = 0
    endT = 0
    while(text.find("HOTELBEDS PTE", endT)!=-1):
        startT = text.find("HOTELBEDS PTE", startT)
        endT = text.find("Please indicate our reference number on each of your invoices.\n--------------------------------------------------------------------------------", startT)
        if endT == -1:
            endT = text.find("Please provide the following services:\n", startT)
        text = text[0:startT].strip()+"\n"+text[endT+len("Please provide the following services:\n \n--------------------------------------------------------------------------------"):len(text)].strip()
    text = text.strip()

    while True:
        if text.find("Page:",0)!=-1:
            text = text[0:text.find("Page:",0)-1].strip() +"\n"+ text[text.find("Page:",0)+11:].strip()
        else:
            break

    #print(text)
    return text
        
# Workbook is created 
#wb = Workbook() 
row = 3

# add_sheet is used to create sheet. 
#sheet1 = wb.add_sheet('Sheet1')

ws.column_dimensions['A'].width = 15
ws.column_dimensions['A'].fill = openpyxl.styles.fills.PatternFill(patternType='solid',fgColor=openpyxl.styles.colors.Color(rgb='FFFF00'))
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 10
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 15
ws.column_dimensions['F'].fill = openpyxl.styles.fills.PatternFill(patternType='solid',fgColor=openpyxl.styles.colors.Color(rgb='FFFF00'))
ws.column_dimensions['G'].width = 15
ws.column_dimensions['H'].width = 15
ws.column_dimensions['I'].width = 15
ws.column_dimensions['J'].width = 15
ws.column_dimensions['K'].width = 32
ws.column_dimensions['L'].width = 22
ws.column_dimensions['M'].width = 15
ws.column_dimensions['N'].width = 40
ws.column_dimensions['O'].width = 42
ws.column_dimensions['P'].width = 70
ws.column_dimensions['Q'].width = 25
ws.column_dimensions['R'].width = 20
ws.column_dimensions['S'].width = 20
ws.column_dimensions['T'].width = 25
ws.column_dimensions['U'].width = 12
ws.column_dimensions['V'].width = 12
ws.column_dimensions['W'].width = 26
ws.column_dimensions['X'].width = 10
ws.column_dimensions['Y'].width = 25
ws.column_dimensions['Z'].width = 50
ws.column_dimensions['AA'].width = 40
ws.column_dimensions['AB'].width = 20
ws.column_dimensions['AC'].width = 40
ws.column_dimensions['AD'].width = 40
ws.column_dimensions['AE'].width = 40
ws.row_dimensions[2].fill = openpyxl.styles.fills.PatternFill(patternType='solid',fgColor=openpyxl.styles.colors.Color(rgb='C0C0C0'))
ws.cell(2, 1, "refNo")
ws.cell(2, 2, "CustomerName")
ws.cell(2, 3, "Agent refNo")
ws.cell(2, 4,"Type")
ws.cell(2, 5,"Booking Date")
ws.cell(2, 6,"Service Id")
ws.cell(2, 7,"Contract")
ws.cell(2, 8,"Service Date")
ws.cell(2, 9,"Service")
ws.cell(2,10,"Modality")
ws.cell(2,11,"Service Description")
ws.cell(2,12,"Modality Description")
ws.cell(2,13,"Rate")
ws.cell(2,14,"PAX")
ws.cell(2,15,"Customer Detail")
ws.cell(2,16,"REMARK")
ws.cell(2,17,"Hotel")
ws.cell(2,18,"Cancel BookingDate")
ws.cell(2,19,"Modification BookingDate")
ws.cell(2,20,"Client Mobile")
ws.cell(2,21,"Arrival")
ws.cell(2,22,"Departure")
ws.cell(2,23,"Commercial description")
ws.cell(2,24,"Serv.Type")
ws.cell(2,25,"Vehicle")
ws.cell(2,26,"From")
ws.cell(2,27,"PickUp Time")
ws.cell(2,28,"PickUp point")
ws.cell(2,29,"Transport")
ws.cell(2,30,"To")
ws.cell(2,31,"OldBooking (Modification)")

for page in folders:
    text2 = convert_pdf_to_txt(page)
    allText = len(text2)
    startText = 0
    indexStart = 0
    indexEndLine = 0
    reference = []
    referenceNum = ""
    referenceName = ""
    typeDetail=""
    mode=""

    while startText <= allText:
        #print(startText)
        #print(allText)
        bookDate=""
        serviceId=""
        contract=""
        serviceDate=""
        service=""
        modality=""
        serviceDes=""
        modalDes=""
        rate=""
        pax=""
        cust=""
        remark=""
        hotel=""
        ModiDate=""
        CancelDate=""
        arrival=""
        depeart=""
        ClientMobile=""
        comDes=""
        serviceType=""
        venhicle=""
        from1=""
        pickuptime=""
        pickuppoint=""
        transport=""
        to1=""
        oldBook = ""
        loop = 0
        indexReference = text2.find("REFERENCE", startText,startText+300)
        if indexReference!=-1: #Page2
            indexLastMinute = text2.find("LAST MINUTE", indexReference+9 , indexReference+50 )
            if indexLastMinute!=-1:
                indexReferenceNum = text2.find("\n", indexLastMinute+11)
                #print(indexReferenceNum)
                indexEndLine = text2.find("\n", indexReferenceNum+1)
                #print(indexEndLine)
            else:
                indexReferenceNum = text2.find("\n", indexReference+9)
                indexEndLine = text2.find("\n", indexReferenceNum+3)
                #print(indexEndLine)
            reference = text2[indexReferenceNum+1:indexEndLine].strip().split()
            #print(reference)
            referenceNum = reference[0]
        if(reference[1]!= "Name:"):
            #print(reference[1])
            mode = "noname"
            referenceName = reference[1]
            for x in range(len(reference)-2):
                if reference[2+x] in ("TRAVEL","MONTROSE","FIVEFLY","CARASOULS","WEBBEDS","SIMPLE","FLIGHT","NOE","MTRAVEL-Main","HOORAY","WAU!"):
                    break
                referenceName += " "+reference[2+x]

            if indexReference!=-1: #Page2
                indexAgencyRef = text2.find("Agency Reference:", indexEndLine)
                if indexAgencyRef !=-1:
                    c = text2.find("\n", indexAgencyRef+1)
                    AgencyReference = text2[indexAgencyRef+17:indexEndLine].strip()
                    #print(indexEndLine)
                else:
                    AgencyReference = ""
                    #print(indexEndLine)

            indexType = text2.find("---", indexEndLine+1,indexEndLine+100)

            if indexType != -1:
                indexEndLine = text2.find("\n", indexType+1)
                #print(indexEndLine)
            typeDetail=""

            if(indexType == -1 or "NEW" in text2[indexType+1:indexType+100]):
                typeDetail="NEW"

                indexBook = text2.find("Creation booking date:", indexEndLine+1)
                if indexBook ==-1:
                    break;
                indexEndLine = text2.find("\n", indexBook+1)
                bookDate = text2[indexBook+22:indexEndLine].strip()

                indexService = text2.find("SERVICE ID", indexEndLine+1)
                indexEndLine = text2.find("\n", indexService+1)
                serviceId = text2[indexService+len("SERVICE ID"):indexEndLine].strip()

                indexStart = text2.find("Contract:", indexEndLine+1)
                indexEndLine = text2.find("\n", indexStart+1)
                contract = text2[indexStart+len("Contract:"):indexEndLine].strip()

                indexStart = text2.find("Service Date", indexEndLine+1)
                indexEndLine = text2.find("\n", indexStart+1)
                serviceDate = text2[indexStart+len("Service Date"):indexEndLine].strip()

                indexStart = text2.find("Service:", indexEndLine+1)
                indexEndLine = text2.find("Modality:", indexStart+1)
                service = text2[indexStart+len("Service:"):indexEndLine-1].strip()

                indexStart = text2.find("Modality:", indexEndLine-1)
                indexEndLine = text2.find("\n", indexStart+1)
                modality = text2[indexStart+len("Modality:"):indexEndLine].strip()

                indexStart = text2.find("Service Description:", indexEndLine+1)
                indexEndLine = text2.find("Modality Description:", indexEndLine+1)
                serviceDes = text2[indexStart+len("Service Description:"):indexEndLine].strip()

                indexStart = indexEndLine
                indexEndLine = text2.find("\n", indexStart+1)
                modalDes = text2[indexStart+len("Modality Description:"):indexEndLine].strip()

                indexStart = text2.find("Rate:", indexEndLine+1)
                indexEndLine = text2.find("\n", indexStart+1)
                rate = text2[indexStart+len("Rate:"):indexEndLine].strip()

                indexStart = text2.find("PASSENGERS:", indexEndLine+1)
                indexEndLine = text2.find("\n", indexStart+1)
                pax = text2[indexStart+len("PASSENGERS:"):indexEndLine].strip()

                indexStart = indexEndLine
                indexEndLine = text2.find("REMARKS:", indexStart-1)
                cust = text2[indexStart:indexEndLine].strip()
                cust1 = cust.splitlines()
                cust=""
                for custs in cust1:
                    cust+=custs.strip() + "\n"
                cust = cust.strip()

                indexStart = text2.find("REMARKS:", indexEndLine)
                indexEndLine = text2.find("Confirmation Number", indexStart-1)
                remark = text2[indexStart+len("REMARKS:"):indexEndLine].strip()

                indexStart = text2.find("of your hotel -", indexStart+1)
                indexEndLine2 = text2.find("\n", indexStart+1)
                hotel = text2[indexStart+len("of your hotel -"):indexEndLine2].strip()

            elif ("CANCELLATION" in text2[indexType+1:indexType+100]):
                typeDetail="CANCEL"

                indexBook = text2.find("Creation booking date:", indexEndLine+1)
                if indexBook == -1:
                    break;
                indexEndLine = text2.find("\n", indexBook+1)
                bookDate = text2[indexBook+len("Creation booking date:"):indexEndLine].strip()

                indexCancel = text2.find("Cancellation booking date:", indexEndLine+1)
                indexEndLine = text2.find("\n", indexCancel+1)
                CancelDate = text2[indexCancel+len("Cancellation booking date:"):indexEndLine].strip()

                indexService = text2.find("SERVICE ID", indexEndLine+1)
                indexEndLine = text2.find("\n", indexService+1)
                serviceId = text2[indexService+len("SERVICE ID"):indexEndLine].strip()

                indexStart = text2.find("Contract:", indexEndLine+1)
                if(indexStart!=-1):
                    indexEndLine = text2.find("\n", indexStart+1)
                    contract = text2[indexStart+len("Contract:"):indexEndLine].strip()
                else:
                    contract=""

                indexStart = text2.find("Service Date", indexEndLine+1)
                if(indexStart!=-1):
                    indexEndLine = text2.find("\n", indexStart+1)
                    serviceDate = text2[indexStart+len("Service Date"):indexEndLine].strip()
                else:
                    serviceDate = ""

                indexStart = text2.find("Service:", indexEndLine+1)
                indexEndLine = text2.find("Modality:", indexStart+1)
                service = text2[indexStart+len("Service:"):indexEndLine-1].strip()

                indexStart = text2.find("Modality:", indexEndLine-1)
                indexEndLine = text2.find("\n", indexStart+1)
                modality = text2[indexStart+len("Modality:"):indexEndLine].strip()

                indexStart = text2.find("Service Description:", indexEndLine+1)
                indexEndLine = text2.find("Modality Description:", indexEndLine+1)
                serviceDes = text2[indexStart+len("Service Description:"):indexEndLine].strip()

                indexStart = indexEndLine
                indexEndLine = text2.find("\n", indexStart+1)
                modalDes = text2[indexStart+len("Modality Description:"):indexEndLine].strip()

                indexStart = text2.find("Rate:", indexEndLine+1)
                indexEndLine = text2.find("\n", indexStart+1)
                rate = text2[indexStart+len("Rate:"):indexEndLine].strip()

                indexStart = text2.find("PASSENGERS:", indexEndLine+1)
                indexEndLine = text2.find("\n", indexStart+1)
                pax = text2[indexStart+len("PASSENGERS:"):indexEndLine].strip()
                #print(indexEndLine)

                indexStart = indexEndLine
                indexEndLine = text2.find("REMARKS:", indexStart-1)
                cust = text2[indexStart:indexEndLine].strip()
                cust1 = cust.splitlines()
                cust=""
                for custs in cust1:
                    cust+=custs.strip() + "\n"
                cust = cust.strip()

                indexStart = text2.find("REMARKS:", indexEndLine)
                indexEndLine = text2.find("Confirmation Number", indexStart-1)
                remark = text2[indexStart+len("REMARKS:"):indexEndLine].strip()

                indexStart = text2.find("of your hotel -", indexStart+1)
                indexEndLine2 = text2.find("\n", indexStart+1)
                hotel = text2[indexStart+len("of your hotel -"):indexEndLine2].strip()

            elif ("MODIFICATION" in text2[indexType+1:indexType+100]):
                typeDetail="MODIFY"

                indexBook = text2.find("Creation booking date:", indexEndLine+1)
                if indexBook == -1:
                    break;
                indexEndLine = text2.find("\n", indexBook+1)
                bookDate = text2[indexBook+len("Creation booking date:"):indexEndLine].strip()

                indexCancel = text2.find("Modification booking date:", indexEndLine+1)
                indexEndLine = text2.find("\n", indexCancel+1)
                ModiDate = text2[indexCancel+len("Modification booking date:"):indexEndLine].strip()

                indexService = text2.find("SERVICE ID", indexEndLine+1)
                indexEndLine = text2.find("\n", indexService+1)
                serviceId = text2[indexService+len("SERVICE ID"):indexEndLine].strip()

                indexStart = text2.find("Contract:", indexEndLine+1)
                indexEndLine = text2.find("\n", indexStart+1)
                contract = text2[indexStart+len("Contract:"):indexEndLine].strip()

                indexStart = text2.find("Service Date", indexEndLine+1)
                indexEndLine = text2.find("\n", indexStart+1)
                serviceDate = text2[indexStart+len("Service Date"):indexEndLine].strip()

                indexStart = text2.find("Service:", indexEndLine+1)
                indexEndLine = text2.find("Modality:", indexStart+1)
                service = text2[indexStart+len("Service:"):indexEndLine-1].strip()

                indexStart = text2.find("Modality:", indexEndLine-1)
                indexEndLine = text2.find("\n", indexStart+1)
                modality = text2[indexStart+len("Modality:"):indexEndLine].strip()

                indexStart = text2.find("Service Description:", indexEndLine+1)
                indexEndLine = text2.find("Modality Description:", indexEndLine+1)
                serviceDes = text2[indexStart+len("Service Description:"):indexEndLine].strip()

                indexStart = indexEndLine
                indexEndLine = text2.find("\n", indexStart+1)
                modalDes = text2[indexStart+len("Modality Description:"):indexEndLine].strip()

                indexStart = text2.find("Rate:", indexEndLine+1)
                indexEndLine = text2.find("\n", indexStart+1)
                rate = text2[indexStart+len("Rate:"):indexEndLine].strip()

                indexStart = text2.find("PASSENGERS:", indexEndLine+1)
                indexEndLine = text2.find("\n", indexStart+1)
                pax = text2[indexStart+len("PASSENGERS:"):indexEndLine].strip()

                indexStart = indexEndLine
                indexEndLine = text2.find("REMARKS:", indexStart-1)
                cust = text2[indexStart:indexEndLine].strip()
                cust1 = cust.splitlines()
                cust=""
                for custs in cust1:
                    cust+=custs.strip() + "\n"
                cust = cust.strip()

                indexStart = text2.find("REMARKS:", indexEndLine)
                indexEndLine = text2.find("Confirmation Number", indexStart-1)
                remark = text2[indexStart+len("REMARKS:"):indexEndLine].strip()

                indexStart = text2.find("of your hotel -", indexStart+1)
                indexEndLine2 = text2.find("\n", indexStart+1)
                hotel = text2[indexStart+len("of your hotel -"):indexEndLine2].strip()
            else:
                break;
        else:
            mode = "withname"
            referenceName = reference[2]
            for x in range(len(reference)-3):
                referenceName += " "+reference[3+x]

            indexAgencyRef = text2.find("TO.Ref.", indexEndLine,indexEndLine+100)
            if indexAgencyRef != -1:
                indexEndLine = text2.find("\n", indexAgencyRef+1)
                AgencyReference = text2[indexAgencyRef+7:indexEndLine].strip()

            indexMobile = text2.find("mobile contact:", indexEndLine ,indexEndLine+100 )
            if indexMobile != -1:
                indexEndLine = text2.find("\n", indexMobile+1)
                ClientMobile = text2[indexMobile+len("mobile contact:"):indexEndLine].strip()
            else:
                ClientMobile=""

            indexType = text2.find("---", indexEndLine+1,indexEndLine+100)
            if indexType != -1:
                indexEndLine = text2.find("\n", indexType+1)
                typeDetail=""

            if(( indexType == -1 and typeDetail=="NEW") or "NEW" in text2[indexType+1:indexType+100]):
                typeDetail="NEW"

                indexArr= text2.find("Arrival:", indexEndLine+1,indexEndLine+200)
                indexEndLineArr = text2.find("\n", indexArr+1)
                if indexArr != -1:
                    arrival = text2[indexArr+len("Arrival:"):indexEndLineArr].strip()
                else:
                    arrival = ""

                indexDe= text2.find("Departure:", indexEndLine+1,indexEndLine+200)
                indexEndLineDe = text2.find("\n", indexDe+1)
                if indexDe != -1:
                    depeart = text2[indexDe+len("Departure:"):indexEndLineDe].strip()
                else:
                    depeart = ""

                if indexArr ==-1 and indexDe == -1: #NotFound Arrival and Departure
                    break

                loop=0
                while True:
                    indexBook = text2.find("Creation booking date:", indexEndLine+1,indexEndLine+100)
                    if indexBook ==-1:
                        loop -=1
                        break
                    indexEndLine = text2.find("\n", indexBook+1)
                    bookDate = text2[indexBook+len("Creation booking date:"):indexEndLine].strip()
                    ws.cell(row+loop,5,bookDate)

                    indexService = text2.find("SERVICE ID:", indexEndLine+1,indexEndLine+100)
                    indexEndLine = text2.find("\n", indexService+1)
                    serviceId = text2[indexService+len("SERVICE ID:"):indexEndLine].strip()
                    ws.cell(row+loop,6,serviceId)

                    indexStart = text2.find("Commercial description:", indexEndLine+1,indexEndLine+100)
                    indexEndLine = text2.find("\n", indexStart+1)
                    comDes = text2[indexStart+len("Commercial description:"):indexEndLine].strip()
                    ws.cell(row+loop,23,comDes)
                    
                    indexStart = text2.find("Serv.Type:", indexEndLine+1,indexEndLine+100)
                    indexEndLine = text2.find("Vehicle:", indexStart+1)
                    serviceType = text2[indexStart+len("Serv.Type:"):indexEndLine-1].strip()
                    ws.cell(row+loop,24,serviceType)

                    indexStart = text2.find("Vehicle:", indexEndLine-1,indexEndLine+100)
                    indexEndLine = text2.find("Paxes:", indexStart+1)
                    venhicle = text2[indexStart+len("Vehicle:"):indexEndLine-1].strip()
                    ws.cell(row+loop,25,venhicle)

                    indexStart = text2.find("Paxes:", indexEndLine-1,indexEndLine+100)
                    indexEndLine = text2.find("\n", indexStart+1)
                    pax = text2[indexStart+len("Paxes:"):indexEndLine].strip()
                    ws.cell(row+loop,14,pax)

                    loop +=1

                indexService = text2.find("From:", indexEndLine+1)
                indexEndLine = text2.find("\n", indexService+1)
                if indexDe != -1:
                    indexEndLineTest = text2.find("\n", indexEndLine+1)
                    indexEndLine = text2.find("\n", indexEndLineTest+1)
                    if text2[indexEndLineTest:indexEndLine].isspace():
                        indexEndLine = indexEndLineTest
                from1 = text2[indexService+len("From:"):indexEndLine].strip()

                indexService = text2.find("Pickup Time:", indexEndLine+1,indexEndLine+100)
                if indexService != -1:
                    indexEndLine = text2.find("hrs", indexService+1)
                    pickuptime = text2[indexService+len("Pickup Time:"):indexEndLine-1].strip()
                else:
                    pickuptime = ""

                indexService = text2.find("Pickup Point:", indexEndLine+1,indexEndLine+100)
                if indexService != -1:
                    indexEndLineP = text2.find("\n", indexService+1)
                    pickuppoint = text2[indexService+len("Pickup Point:"):indexEndLineP].strip()
                else:
                    pickuppoint = ""

                indexService = text2.find("Transport:", indexEndLineP+1)
                indexEndLine1 = text2.find("\n", indexService+1)
                transport = text2[indexService+len("Transport:"):indexEndLine1].strip()

                indexService = text2.find("To: ", indexEndLineP+1)
                indexEndLine = text2.find("\n", indexService+1)
                if indexArr != -1:
                    indexEndLineTest = text2.find("\n", indexEndLine+1)
                    indexEndLine = text2.find("\n", indexEndLineTest+1)
                    if text2[indexEndLineTest:indexEndLine].isspace():
                        indexEndLine = indexEndLineTest
                to1 = text2[indexService+len("To: "):indexEndLine].strip()

                if indexEndLine1 > indexEndLine:
                    indexEndLine = indexEndLine1

            elif (( indexType == -1 and typeDetail=="CANCEL") or"CANCELLATION" in text2[indexType+1:indexType+100]):
                typeDetail="CANCEL"

                indexArr= text2.find("Arrival:", indexEndLine+1,indexEndLine+200)
                indexEndLineArr = text2.find("\n", indexArr+1)
                if indexArr != -1:
                    arrival = text2[indexArr+len("Arrival:"):indexEndLineArr].strip()
                else:
                    arrival=""

                indexDe= text2.find("Departure:", indexEndLine+1,indexEndLine+200)
                indexEndLineDe = text2.find("\n", indexDe+1)
                if indexDe != -1:
                    depeart = text2[indexDe+len("Departure:"):indexEndLineDe].strip()
                else:
                    depeart=""

                if indexArr ==-1 and indexDe == -1: #NotFound Arrival and Departure
                    break

                loop=0
                while True:
                    indexBook = text2.find("Creation booking date:", indexEndLine+1,indexEndLine+100)
                    if indexBook ==-1:
                        loop -=1
                        break;
                    indexEndLine = text2.find("\n", indexBook+1)
                    bookDate = text2[indexBook+len("Creation booking date:"):indexEndLine].strip()
                    ws.cell(row+loop,5,bookDate)

                    indexCancel = text2.find("Cancellation booking date:", indexEndLine+1,indexEndLine+100)
                    indexEndLine = text2.find("\n", indexCancel+1)
                    CancelDate = text2[indexCancel+len("Cancellation booking date:"):indexEndLine].strip()
                    ws.cell(row+loop,18,CancelDate)

                    indexService = text2.find("SERVICE ID:", indexEndLine+1,indexEndLine+100)
                    indexEndLine = text2.find("\n", indexService+1)
                    serviceId = text2[indexService+len("SERVICE ID:"):indexEndLine].strip()
                    ws.cell(row+loop,6,serviceId)

                    indexStart = text2.find("Commercial description:", indexEndLine+1,indexEndLine+100)
                    indexEndLine = text2.find("\n", indexStart+1)
                    comDes = text2[indexStart+len("Commercial description:"):indexEndLine].strip()
                    ws.cell(row+loop,23,comDes)
                    
                    indexStart = text2.find("Serv.Type:", indexEndLine+1,indexEndLine+100)
                    indexEndLine = text2.find("Vehicle:", indexStart+1)
                    serviceType = text2[indexStart+len("Serv.Type:"):indexEndLine-1].strip()
                    ws.cell(row+loop,24,serviceType)

                    indexStart = text2.find("Vehicle:", indexEndLine-1,indexEndLine+100)
                    indexEndLine = text2.find("Paxes:", indexStart+1)
                    venhicle = text2[indexStart+len("Vehicle:"):indexEndLine-1].strip()
                    ws.cell(row+loop,25,venhicle)

                    indexStart = text2.find("Paxes:", indexEndLine-1,indexEndLine+100)
                    indexEndLine = text2.find("\n", indexStart+1)
                    pax = text2[indexStart+len("Paxes:"):indexEndLine].strip()
                    ws.cell(row+loop,14,pax)

                    loop +=1

                indexService = text2.find("From:", indexEndLine+1)
                indexEndLine = text2.find("\n", indexService+1)
                if indexDe != -1:
                    indexEndLineTest = text2.find("\n", indexEndLine+1)
                    indexEndLine = text2.find("\n", indexEndLineTest+1)
                    if text2[indexEndLineTest:indexEndLine].isspace():
                        indexEndLine = indexEndLineTest
                from1 = text2[indexService+len("From:"):indexEndLine].strip()

                indexService = text2.find("Pickup Time:", indexEndLine+1,indexEndLine+100)
                if indexService != -1:
                    indexEndLine = text2.find("hrs", indexService+1)
                    pickuptime = text2[indexService+len("Pickup Time:"):indexEndLine-1].strip()
                else:
                    pickuptime=""

                indexService = text2.find("Pickup Point:", indexEndLine+1,indexEndLine+100)
                if indexService != -1:
                    indexEndLineP = text2.find("\n", indexService+1)
                    pickuppoint = text2[indexService+len("Pickup Point:"):indexEndLineP].strip()
                else:
                    pickuppoint =""

                indexService = text2.find("To: ", indexEndLineP+1)
                indexEndLine = text2.find("\n", indexService+1)
                if indexArr != -1:
                    indexEndLineTest = text2.find("\n", indexEndLine+1)
                    indexEndLine = text2.find("\n", indexEndLineTest+1)
                    if text2[indexEndLineTest:indexEndLine].isspace():
                        indexEndLine = indexEndLineTest
                to1 = text2[indexService+len("To: "):indexEndLine].strip()

                indexService = text2.find("Transport:", indexEndLineP+1)
                indexEndLine1 = text2.find("\n", indexService+1)
                transport = text2[indexService+len("Transport:"):indexEndLine1].strip()

                if indexEndLine1 > indexEndLine:
                    indexEndLine = indexEndLine1

            elif (( indexType == -1 and typeDetail=="MODIFY") or "MODIFICATION" in text2[indexType+1:indexType+100]):
                typeDetail="MODIFY"

                indexArr= text2.find("Arrival:", indexEndLine+1,indexEndLine+200)
                indexEndLineArr = text2.find("\n", indexArr+1)
                if indexArr != -1:
                    arrival = text2[indexArr+len("Arrival:"):indexEndLineArr].strip()
                else:
                    arrival=""

                indexDe= text2.find("Departure:", indexEndLine+1,indexEndLine+200)
                indexEndLineDe = text2.find("\n", indexDe+1)
                if indexDe != -1:
                    depeart = text2[indexDe+len("Departure:"):indexEndLineDe].strip()
                else:
                    depeart=""

                if indexArr ==-1 and indexDe == -1: #NotFound Arrival and Departure
                    break

                loop=0
                while True:
                    indexBook = text2.find("Creation booking date:", indexEndLine+1,indexEndLine+100)
                    if indexBook ==-1:
                        loop -=1
                        break;
                    indexEndLine = text2.find("\n", indexBook+1)
                    bookDate = text2[indexBook+len("Creation booking date:"):indexEndLine].strip()
                    ws.cell(row+loop,5,bookDate)

                    indexCancel = text2.find("Modification booking date:", indexEndLine+1,indexEndLine+100)
                    indexEndLine = text2.find("\n", indexCancel+1)
                    ModiDate = text2[indexCancel+len("Modification booking date:"):indexEndLine].strip()
                    ws.cell(row+loop,19,ModiDate)

                    indexService = text2.find("SERVICE ID:", indexEndLine+1,indexEndLine+100)
                    indexEndLine = text2.find("\n", indexService+1)
                    serviceId = text2[indexService+len("SERVICE ID:"):indexEndLine].strip()
                    ws.cell(row+loop,6,serviceId)

                    indexStart = text2.find("Commercial description:", indexEndLine+1,indexEndLine+100)
                    indexEndLine = text2.find("\n", indexStart+1)
                    comDes = text2[indexStart+len("Commercial description:"):indexEndLine].strip()
                    ws.cell(row+loop,23,comDes)
                    
                    indexStart = text2.find("Serv.Type:", indexEndLine+1,indexEndLine+100)
                    indexEndLine = text2.find("Vehicle:", indexStart+1)
                    serviceType = text2[indexStart+len("Serv.Type:"):indexEndLine-1].strip()
                    ws.cell(row+loop,24,serviceType)

                    indexStart = text2.find("Vehicle:", indexEndLine-1,indexEndLine+100)
                    indexEndLine = text2.find("Paxes:", indexStart+1)
                    venhicle = text2[indexStart+len("Vehicle:"):indexEndLine-1].strip()
                    ws.cell(row+loop,25,venhicle)

                    indexStart = text2.find("Paxes:", indexEndLine-1,indexEndLine+100)
                    indexEndLine = text2.find("\n", indexStart+1)
                    pax = text2[indexStart+len("Paxes:"):indexEndLine].strip()
                    ws.cell(row+loop,14,pax)

                    loop+=1


                indexService = text2.find("From:", indexEndLine+1)
                indexEndLine = text2.find("\n", indexService+1)
                if indexDe != -1:
                    indexEndLineTest = text2.find("\n", indexEndLine+1)
                    indexEndLine = text2.find("\n", indexEndLineTest+1)
                    if text2[indexEndLineTest:indexEndLine].isspace():
                        indexEndLine = indexEndLineTest
                from1 = text2[indexService+len("From:"):indexEndLine].strip()

                indexService = text2.find("Pickup Time:", indexEndLine+1,indexEndLine+100)
                if indexService != -1:
                    indexEndLine = text2.find("hrs", indexService+1)
                    pickuptime = text2[indexService+len("Pickup Time:"):indexEndLine-1].strip()
                else:
                    pickuptime=""

                indexService = text2.find("Pickup Point:", indexEndLine+1,indexEndLine+100)
                if indexService != -1:
                    indexEndLineP = text2.find("\n", indexService+1)
                    pickuppoint = text2[indexService+len("Pickup Point:"):indexEndLineP].strip()
                else:
                    pickuppoint = ""

                indexService = text2.find("To: ", indexEndLineP+1)
                indexEndLine = text2.find("\n", indexService+1)
                if indexArr != -1:
                    indexEndLineTest = text2.find("\n", indexEndLine+1)
                    indexEndLine = text2.find("\n", indexEndLineTest+1)
                    if text2[indexEndLineTest:indexEndLine].isspace():
                        indexEndLine = indexEndLineTest
                to1 = text2[indexService+len("To: "):indexEndLine].strip()

                indexService = text2.find("Transport:", indexEndLineP+1)
                indexEndLine1 = text2.find("\n", indexService+1)
                transport = text2[indexService+len("Transport:"):indexEndLine1].strip()

                if indexEndLine1 > indexEndLine:
                    indexEndLine = indexEndLine1

                indexService = text2.find("Old booking:", indexEndLine+1)
                indexEndLine = text2.find("Transport:", indexService+1)
                indexEndLine = text2.find("To:", indexEndLine+1)
                indexEndLine = text2.find("\n", indexEndLine+1)
                oldBook = text2[indexService+len("Old booking:"):indexEndLine].strip()
            else:
                break

        
        if mode != "withname" :
            ws.cell(row,5,bookDate)
            ws.cell(row,6,serviceId)
            ws.cell(row,14,pax)
            ws.cell(row,23,comDes)
            ws.cell(row,24,serviceType)
            ws.cell(row,25,venhicle)
            ws.cell(row,19,ModiDate)
            ws.cell(row,18,CancelDate)

        saveloop = loop
        while loop >-1:
            ws.cell(row+loop,1,referenceNum)
            ws.cell(row+loop,2,referenceName) 
            ws.cell(row+loop,3,AgencyReference)
            ws.cell(row+loop,4,typeDetail)
            ws.cell(row+loop,7,contract)
            ws.cell(row+loop,8,serviceDate)
            ws.cell(row+loop,9,service)
            ws.cell(row+loop,10,modality)
            ws.cell(row+loop,11,serviceDes)
            ws.cell(row+loop,12,modalDes)
            ws.cell(row+loop,13,rate)
            ws.cell(row+loop,15,cust)
            ws.cell(row+loop,16,remark)
            ws.cell(row+loop,17,hotel)
            ws.cell(row+loop,20,ClientMobile)
            ws.cell(row+loop,21,arrival)
            ws.cell(row+loop,22,depeart)
            ws.cell(row+loop,26,from1)
            ws.cell(row+loop,27,pickuptime)
            ws.cell(row+loop,28,pickuppoint)
            ws.cell(row+loop,29,transport)
            ws.cell(row+loop,30,to1)
            ws.cell(row+loop,31,oldBook)
            loop -=1

        startText = indexEndLine+1
        #print(indexEndLine)
        #print("********************")
        #print(text2[startText:startText+400])
        #print(row)
        row +=1+saveloop
wb.save('DatabaseHB.xlsm') 